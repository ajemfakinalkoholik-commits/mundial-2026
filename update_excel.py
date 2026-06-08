import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
import re

def get_team_info(country_name):
    mapping = {
        'Mexico': ('MEX', 'mx'), 'South Africa': ('RSA', 'za'), 'South Korea': ('KOR', 'kr'), 'Czech Republic': ('CZE', 'cz'),
        'Canada': ('CAN', 'ca'), 'Bosnia and Herzegovina': ('BIH', 'ba'), 'Qatar': ('QAT', 'qa'), 'Switzerland': ('SUI', 'ch'),
        'Brazil': ('BRA', 'br'), 'Morocco': ('MAR', 'ma'), 'Haiti': ('HAI', 'ht'), 'Scotland': ('SCO', 'gb-sct'),
        'United States': ('USA', 'us'), 'Paraguay': ('PAR', 'py'), 'Australia': ('AUS', 'au'), 'Kosovo': ('KOS', 'xk'),
        'Germany': ('GER', 'de'), 'Curaçao': ('CUW', 'cw'), 'Japan': ('JPN', 'jp'), 'Nigeria': ('NGA', 'ng'),
        'Netherlands': ('NED', 'nl'), 'Ecuador': ('ECU', 'ec'), 'Saudi Arabia': ('KSA', 'sa'), 'Panama': ('PAN', 'pa'),
        'Belgium': ('BEL', 'be'), 'Egypt': ('EGY', 'eg'), 'Uruguay': ('URU', 'uy'), 'New Zealand': ('NZL', 'nz'),
        'Spain': ('ESP', 'es'), 'Cape Verde': ('CPV', 'cv'), 'Serbia': ('SRB', 'rs'), 'Mali': ('MLI', 'ml'),
        'France': ('FRA', 'fr'), 'Senegal': ('SEN', 'sn'), 'Iran': ('IRN', 'ir'), 'Costa Rica': ('CRC', 'cr'),
        'Argentina': ('ARG', 'ar'), 'Algeria': ('ALG', 'dz'), 'Colombia': ('COL', 'co'), 'Peru': ('PER', 'pe'),
        'Portugal': ('POR', 'pt'), 'DR Congo': ('COD', 'cd'), 'Wales': ('WAL', 'gb-wls'), 'Cameroon': ('CMR', 'cm'),
        'England': ('ENG', 'gb-eng'), 'Croatia': ('CRO', 'hr'), 'Ivory Coast': ('CIV', 'ci'), 'Jamaica': ('JAM', 'jm'),
        'Italy': ('ITA', 'it'), 'Ukraine': ('UKR', 'ua'), 'Turkey': ('TUR', 'tr'), 'Tunisia': ('TUN', 'tn'),
        'Chile': ('CHI', 'cl'), 'Sweden': ('SWE', 'se'), 'Poland': ('POL', 'pl'), 'Denmark': ('DEN', 'dk')
    }
    val = mapping.get(country_name, (country_name[:3].upper(), 'xx'))
    return val[0], f"https://flagcdn.com/w40/{val[1]}.png"

def parse_schedule():
    import urllib.request
    url = 'https://en.wikipedia.org/wiki/2026_FIFA_World_Cup'
    req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/110.0'})
    response_text = urllib.request.urlopen(req).read().decode('utf-8')
    soup = BeautifulSoup(response_text, 'html.parser')

    groups = {}
    matches = {}

    for h3 in soup.find_all('h3'):
        if h3.text.startswith('Group '):
            group_name = h3.text.split('[')[0].strip()
            table = h3.find_next('table', class_='wikitable')
            if table:
                teams = []
                for tr in table.find_all('tr')[1:]:
                    th = tr.find('th')
                    if th:
                        team = th.text.strip().replace(' (H)', '').replace(' vte', '')
                        if team and team != 'Pos':
                            teams.append(team)
                if teams:
                    groups[group_name] = teams
                    matches[group_name] = []

    # Get matches for groups
    for div in soup.find_all('div', class_='footballbox'):
        try:
            team1 = div.find('th', class_='fhome').text.strip()
            team2 = div.find('th', class_='faway').text.strip()
            
            group_of_match = None
            for g_name, g_teams in groups.items():
                if team1 in g_teams or team2 in g_teams:
                    group_of_match = g_name
                    break
            
            if not group_of_match:
                continue

            date_elem = div.find('div', class_='fdate')
            time_elem = div.find('div', class_='ftime')
            
            date_str = date_elem.text.strip() if date_elem else ""
            time_str = time_elem.text.strip() if time_elem else ""
            
            m = re.search(r'\((2026-\d{2}-\d{2})\)', date_str)
            iso_date = m.group(1) if m else "2026-06-XX"
            
            polish_time = ""
            if time_str:
                t_m = re.search(r'(\d{1,2}:\d{2})', time_str)
                tz_m = re.search(r'(EDT|CDT|MDT|PDT)', time_str)
                if t_m:
                    base_t = t_m.group(1)
                    hr, mn = map(int, base_t.split(':'))
                    offset = 0
                    if tz_m:
                        tz = tz_m.group(1)
                        if tz == 'EDT': offset = 6
                        elif tz == 'CDT': offset = 7
                        elif tz == 'MDT': offset = 8
                        elif tz == 'PDT': offset = 9
                    else:
                        offset = 7
                    
                    hr_pl = (hr + offset) % 24
                    polish_time = f"{hr_pl:02d}:{mn:02d} (Czas PL)"
            else:
                polish_time = "Godzina nieznana"

            matches[group_of_match].append({
                'date': f"{iso_date} | {polish_time}",
                'team1': team1,
                'team2': team2
            })

        except Exception as e:
            pass

    return groups, matches

def create_excel_v2(groups, matches):
    if len(groups) != 12:
        raise ValueError(f'Expected 12 groups, got {len(groups)}')
    wb = openpyxl.Workbook()
    
    DARK_BG = "1E1E1E"
    LIGHT_TEXT = "FFFFFF"
    ACCENT_GREEN = "4CAF50"
    ACCENT_RED = "F44336"
    ACCENT_YELLOW = "FFEB3B"
    BEST_BET_COLOR = "0000FF" # Pure Blue for highest contrast
    GRAY_BG = "E0E0E0"
    
    thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- Arkusz 1: Instrukcja i Podsumowanie ---
    ws_main = wb.active
    ws_main.title = "Podsumowanie i Instrukcja"
    ws_main.protection.sheet = True
    ws_main.sheet_view.showGridLines = False
    
    ws_main.column_dimensions['A'].width = 30
    ws_main.column_dimensions['B'].width = 15
    ws_main.column_dimensions['C'].width = 5
    ws_main.column_dimensions['D'].width = 25
    ws_main.column_dimensions['E'].width = 15

    # World Cup Logo
    ws_main.merge_cells('A1:B3')
    ws_main['A1'] = '=IMAGE("https://upload.wikimedia.org/wikipedia/en/thumb/4/43/2026_FIFA_World_Cup.svg/512px-2026_FIFA_World_Cup.svg.png")'
    ws_main['A1'].alignment = Alignment(horizontal="center", vertical="center")

    # Instrukcja
    ws_main['A5'] = "Zasady Punktacji"
    ws_main['A5'].font = Font(bold=True, size=14)
    ws_main['A6'] = "Dokładny wynik (np. 2:1 i było 2:1):"
    ws_main['B6'] = 10
    ws_main['A7'] = "Dobry rezultat (np. 1X2):"
    ws_main['B7'] = 2

    ws_main['A9'] = "Lista Graczy"
    ws_main['A9'].font = Font(bold=True, size=14)
    for i in range(8):
        ws_main[f'A{10+i}'] = f"Gracz {i+1}"
        ws_main[f'A{10+i}'].fill = PatternFill(start_color=GRAY_BG, end_color=GRAY_BG, fill_type="solid")

    # Podsumowanie Wyników
    ws_main['D5'] = "Klasyfikacja Typerów"
    ws_main['D5'].font = Font(bold=True, size=14)
    ws_main['D6'] = "Gracz"
    ws_main['E6'] = "Suma Punktów"
    ws_main['D6'].fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
    ws_main['D6'].font = Font(color=LIGHT_TEXT, bold=True)
    ws_main['E6'].fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
    ws_main['E6'].font = Font(color=LIGHT_TEXT, bold=True)

    group_names = sorted(list(groups.keys()))
    
    # --- Grupy ---
    for g_idx, g_name in enumerate(group_names):
        ws_g = wb.create_sheet(g_name.replace('Group', 'Grupa'))
        ws_g.sheet_view.showGridLines = False
        ws_g.protection.sheet = True
        g_teams = groups[g_name]
        g_matches = matches.get(g_name, [])
        
        # Columns config
        ws_g.column_dimensions['A'].width = 2
        ws_g.column_dimensions['B'].width = 6   # Flag A
        ws_g.column_dimensions['C'].width = 18  # Team A
        ws_g.column_dimensions['D'].width = 5   # Score A
        ws_g.column_dimensions['E'].width = 3   # :
        ws_g.column_dimensions['F'].width = 5   # Score B
        ws_g.column_dimensions['G'].width = 18  # Team B
        ws_g.column_dimensions['H'].width = 6   # Flag B
        
        ws_g.column_dimensions['I'].width = 3   # Separator
        
        for col_idx in [10, 11, 12, 13]:
            ws_g.column_dimensions[get_column_letter(col_idx)].hidden = True
        
        # Users 1-8 cols
        c_u = 14 # N
        for u in range(8):
            c1 = get_column_letter(c_u)
            c2 = get_column_letter(c_u+1)
            c3 = get_column_letter(c_u+2)
            c4 = get_column_letter(c_u+3) # hidden pts
            c5 = get_column_letter(c_u+4) # Separator
            
            ws_g.column_dimensions[c1].width = 5.42
            ws_g.column_dimensions[c2].width = 5.42
            ws_g.column_dimensions[c3].width = 5.42
            ws_g.column_dimensions[c4].width = 0 
            ws_g.column_dimensions[c4].hidden = True
            ws_g.column_dimensions[c5].width = 3
            
            c_u += 5
            
        row = 2
        match_rows = []
        
        for m_idx, m in enumerate(g_matches):
            # Row 1: Date & Time, User Names
            ws_g.merge_cells(f'B{row}:H{row}')
            ws_g[f'B{row}'] = m['date']
            ws_g[f'B{row}'].alignment = Alignment(horizontal="center")
            ws_g[f'B{row}'].font = Font(bold=True, color=LIGHT_TEXT)
            ws_g[f'B{row}'].fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
            
            pts_cells = []
            
            c_u = 14
            for u in range(8):
                c_start = get_column_letter(c_u)
                c_end = get_column_letter(c_u+2)
                ws_g.merge_cells(f'{c_start}{row}:{c_end}{row}')
                name_cell = ws_g[f'{c_start}{row}']
                name_cell.value = f"='Podsumowanie i Instrukcja'!A{10+u}"
                name_cell.alignment = Alignment(horizontal="center")
                name_cell.font = Font(bold=True)
                name_cell.border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'))
                
                pts_col = get_column_letter(c_u+3)
                pts_cells.append(f"${pts_col}${row+1}")
                c_u += 5
                
            row += 1
            
            # Row 2: Match actual data and user predictions
            act_a = f"D{row}"
            act_b = f"F{row}"
            match_rows.append((act_a, act_b, m['team1'], m['team2']))
            
            team1_code, team1_flag = get_team_info(m['team1'])
            team2_code, team2_flag = get_team_info(m['team2'])
            
            ws_g[f'B{row}'] = f'=IMAGE("{team1_flag}")'
            ws_g[f'C{row}'] = team1_code
            ws_g[f'D{row}'] = ""
            ws_g[f'E{row}'] = ":"
            ws_g[f'F{row}'] = ""
            ws_g[f'G{row}'] = team2_code
            ws_g[f'H{row}'] = f'=IMAGE("{team2_flag}")'
            
            ws_g[f'C{row}'].alignment = Alignment(horizontal="right")
            ws_g[f'G{row}'].alignment = Alignment(horizontal="left")
            for col in ['B','D','E','F','H']:
                ws_g[f'{col}{row}'].alignment = Alignment(horizontal="center")
                
            ws_g[act_a].border = thin_border
            ws_g[act_b].border = thin_border
            ws_g[act_a].protection = Protection(locked=False)
            ws_g[act_b].protection = Protection(locked=False)
            
            # Conditional Formatting for Team Names (Win/Loss/Draw)
            rule_win_a = FormulaRule(formula=[f'AND(ISNUMBER({act_a}), ISNUMBER({act_b}), {act_a}>{act_b})'], stopIfTrue=True, fill=PatternFill(start_color=ACCENT_GREEN, end_color=ACCENT_GREEN, fill_type="solid"))
            rule_loss_a = FormulaRule(formula=[f'AND(ISNUMBER({act_a}), ISNUMBER({act_b}), {act_a}<{act_b})'], stopIfTrue=True, fill=PatternFill(start_color=ACCENT_RED, end_color=ACCENT_RED, fill_type="solid"))
            rule_draw_a = FormulaRule(formula=[f'AND(ISNUMBER({act_a}), ISNUMBER({act_b}), {act_a}={act_b})'], stopIfTrue=True, fill=PatternFill(start_color=ACCENT_YELLOW, end_color=ACCENT_YELLOW, fill_type="solid"))
            
            ws_g.conditional_formatting.add(f'C{row}', rule_win_a)
            ws_g.conditional_formatting.add(f'C{row}', rule_loss_a)
            ws_g.conditional_formatting.add(f'C{row}', rule_draw_a)
            
            rule_win_b = FormulaRule(formula=[f'AND(ISNUMBER({act_a}), ISNUMBER({act_b}), {act_b}>{act_a})'], stopIfTrue=True, fill=PatternFill(start_color=ACCENT_GREEN, end_color=ACCENT_GREEN, fill_type="solid"))
            rule_loss_b = FormulaRule(formula=[f'AND(ISNUMBER({act_a}), ISNUMBER({act_b}), {act_b}<{act_a})'], stopIfTrue=True, fill=PatternFill(start_color=ACCENT_RED, end_color=ACCENT_RED, fill_type="solid"))
            rule_draw_b = FormulaRule(formula=[f'AND(ISNUMBER({act_a}), ISNUMBER({act_b}), {act_b}={act_a})'], stopIfTrue=True, fill=PatternFill(start_color=ACCENT_YELLOW, end_color=ACCENT_YELLOW, fill_type="solid"))

            ws_g.conditional_formatting.add(f'G{row}', rule_win_b)
            ws_g.conditional_formatting.add(f'G{row}', rule_loss_b)
            ws_g.conditional_formatting.add(f'G{row}', rule_draw_b)
            
            # User inputs
            c_u = 14
            for u in range(8):
                u_a = get_column_letter(c_u) + str(row)
                u_colon = get_column_letter(c_u+1) + str(row)
                u_b = get_column_letter(c_u+2) + str(row)
                u_pts = get_column_letter(c_u+3) + str(row)
                
                ws_g[u_colon] = ":"
                ws_g[u_colon].alignment = Alignment(horizontal="center")
                ws_g[u_a].border = Border(left=Side(style='medium'), top=Side(style='thin'), bottom=Side(style='medium'))
                ws_g[u_b].border = Border(right=Side(style='medium'), top=Side(style='thin'), bottom=Side(style='medium'))
                ws_g[u_colon].border = Border(bottom=Side(style='medium'))
                ws_g[u_a].alignment = Alignment(horizontal="center")
                ws_g[u_b].alignment = Alignment(horizontal="center")
                ws_g[u_a].protection = Protection(locked=False)
                ws_g[u_b].protection = Protection(locked=False)
                
                # Points formula
                f_str = f'=IF(AND(ISNUMBER({u_a}), ISNUMBER({u_b}), ISNUMBER({act_a}), ISNUMBER({act_b})), IF(AND({u_a}={act_a}, {u_b}={act_b}), \'Podsumowanie i Instrukcja\'!$B$6, IF(SIGN({u_a}-{u_b})=SIGN({act_a}-{act_b}), \'Podsumowanie i Instrukcja\'!$B$7, 0)), 0)'
                ws_g[u_pts] = f_str
                
                # Highlight logic for User Name (Row-1)
                name_cell_addr = f'{get_column_letter(c_u)}{row-1}:{get_column_letter(c_u+2)}{row-1}'
                
                # Exact score rule
                rule_exact = FormulaRule(formula=[f'AND(ISNUMBER({u_pts}), {u_pts}=INDIRECT("\'Podsumowanie i Instrukcja\'!$B$6"))'], stopIfTrue=False, fill=PatternFill(start_color=ACCENT_GREEN, end_color=ACCENT_GREEN, fill_type="solid"))
                
                # Correct result rule
                rule_res = FormulaRule(formula=[f'AND(ISNUMBER({u_pts}), {u_pts}=INDIRECT("\'Podsumowanie i Instrukcja\'!$B$7"))'], stopIfTrue=False, fill=PatternFill(start_color=ACCENT_YELLOW, end_color=ACCENT_YELLOW, fill_type="solid"))
                
                # Wrong result rule (match played but 0 points)
                rule_wrong = FormulaRule(formula=[f'AND(ISNUMBER({act_a}), {u_pts}=0)'], stopIfTrue=False, fill=PatternFill(start_color=ACCENT_RED, end_color=ACCENT_RED, fill_type="solid"))
                
                ws_g.conditional_formatting.add(name_cell_addr, rule_exact)
                ws_g.conditional_formatting.add(name_cell_addr, rule_res)
                ws_g.conditional_formatting.add(name_cell_addr, rule_wrong)
                
                c_u += 5
            
            row += 3
            
        # --- TABELE GRUPOWE ---
        row += 2
        calc_start_row = row
        
        # Hide calculation columns (CA onwards)
        for col_i in range(79, 130):
            ws_g.column_dimensions[get_column_letter(col_i)].hidden = True
            
        # Official Table Calcs
        ws_g[f'CA{calc_start_row}'] = "Team"
        ws_g[f'CB{calc_start_row}'] = "Pkt"
        ws_g[f'CC{calc_start_row}'] = "B+"
        ws_g[f'CD{calc_start_row}'] = "B-"
        ws_g[f'CE{calc_start_row}'] = "B+/-"
        
        for idx, t in enumerate(g_teams):
            r = calc_start_row + 1 + idx
            t_code, _ = get_team_info(t)
            ws_g[f'CA{r}'] = t_code
            
            p_f, w_f, d_f, l_f, gf_f, ga_f = [], [], [], [], [], []
            
            for mr in match_rows:
                act_a, act_b, t_a, t_b = mr[0], mr[1], mr[2], mr[3]
                
                if t_a == t:
                    w_f.append(f"IF(AND(ISNUMBER({act_a}), ISNUMBER({act_b}), {act_a}>{act_b}), 1, 0)")
                    d_f.append(f"IF(AND(ISNUMBER({act_a}), ISNUMBER({act_b}), {act_a}={act_b}), 1, 0)")
                    gf_f.append(f"IF(ISNUMBER({act_a}), {act_a}, 0)")
                    ga_f.append(f"IF(ISNUMBER({act_b}), {act_b}, 0)")
                elif t_b == t:
                    w_f.append(f"IF(AND(ISNUMBER({act_b}), ISNUMBER({act_a}), {act_b}>{act_a}), 1, 0)")
                    d_f.append(f"IF(AND(ISNUMBER({act_b}), ISNUMBER({act_a}), {act_b}={act_a}), 1, 0)")
                    gf_f.append(f"IF(ISNUMBER({act_b}), {act_b}, 0)")
                    ga_f.append(f"IF(ISNUMBER({act_a}), {act_a}, 0)")
            
            w_str = '+'.join(w_f) if w_f else '0'
            d_str = '+'.join(d_f) if d_f else '0'
            ws_g[f'CB{r}'] = f"=(({w_str})*3)+(({d_str})*1)"
            ws_g[f'CC{r}'] = f"={'+'.join(gf_f) if gf_f else '0'}"
            ws_g[f'CD{r}'] = f"={'+'.join(ga_f) if ga_f else '0'}"
            ws_g[f'CE{r}'] = f"=CC{r}-CD{r}"

        # Official Table UI
        ws_g.merge_cells(f'B{calc_start_row}:F{calc_start_row}')
        ws_g[f'B{calc_start_row}'] = "Oficjalna Tabela"
        ws_g[f'B{calc_start_row}'].font = Font(bold=True, color=LIGHT_TEXT)
        ws_g[f'B{calc_start_row}'].fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
        ws_g[f'B{calc_start_row}'].alignment = Alignment(horizontal="center")
        
        headers = ["Poz", "Kraj", "Pkt", "B+", "B-"]
        for idx, h in enumerate(headers):
            c = get_column_letter(2+idx)
            ws_g[f'{c}{calc_start_row+1}'] = h
            ws_g[f'{c}{calc_start_row+1}'].font = Font(bold=True)
            ws_g[f'{c}{calc_start_row+1}'].alignment = Alignment(horizontal="center")
        
        # SORT(CA1:CD4, CB1:CB4, FALSE, CE1:CE4, FALSE, CC1:CC4, FALSE)
        sort_f = f"=SORT(CA{calc_start_row+1}:CD{calc_start_row+4}, CB{calc_start_row+1}:CB{calc_start_row+4}, FALSE, CE{calc_start_row+1}:CE{calc_start_row+4}, FALSE, CC{calc_start_row+1}:CC{calc_start_row+4}, FALSE)"
        ws_g[f'C{calc_start_row+2}'] = sort_f
        
        for idx in range(4):
            ws_g[f'B{calc_start_row+2+idx}'] = idx + 1
            ws_g[f'B{calc_start_row+2+idx}'].alignment = Alignment(horizontal="center")
            
        # User Tables
        c_u = 14
        for u in range(8):
            start_col_letter = get_column_letter(c_u)
            end_col_letter = get_column_letter(c_u+2)
            
            ws_g.merge_cells(f'{start_col_letter}{calc_start_row}:{end_col_letter}{calc_start_row}')
            c_header = ws_g[f'{start_col_letter}{calc_start_row}']
            c_header.value = f"='Podsumowanie i Instrukcja'!A{10+u}"
            c_header.font = Font(bold=True)
            c_header.alignment = Alignment(horizontal="center")
            c_header.fill = PatternFill(start_color=GRAY_BG, end_color=GRAY_BG, fill_type="solid")
            
            ws_g[f'{start_col_letter}{calc_start_row+1}'] = "Poz"
            ws_g[f'{get_column_letter(c_u+1)}{calc_start_row+1}'] = "Kraj"
            ws_g[f'{get_column_letter(c_u+2)}{calc_start_row+1}'] = "Pkt"
            
            u_calc_start_col = 85 + (u*2) # CG, CI, CK etc
            ws_g[f'{get_column_letter(u_calc_start_col)}{calc_start_row}'] = "Team"
            ws_g[f'{get_column_letter(u_calc_start_col+1)}{calc_start_row}'] = "Pts"
            
            for idx, t in enumerate(g_teams):
                r = calc_start_row + 1 + idx
                t_code, _ = get_team_info(t)
                ws_g[f'{get_column_letter(u_calc_start_col)}{r}'] = t_code
                
                pts_f = []
                row_m = 3
                for mr in match_rows:
                    t_a, t_b = mr[2], mr[3]
                    u_a = get_column_letter(c_u) + str(row_m)
                    u_b = get_column_letter(c_u+2) + str(row_m)
                    
                    if t_a == t:
                        pts_f.append(f"IF(AND(ISNUMBER({u_a}), ISNUMBER({u_b}), {u_a}>{u_b}), 3, IF(AND(ISNUMBER({u_a}), ISNUMBER({u_b}), {u_a}={u_b}), 1, 0))")
                    elif t_b == t:
                        pts_f.append(f"IF(AND(ISNUMBER({u_b}), ISNUMBER({u_a}), {u_b}>{u_a}), 3, IF(AND(ISNUMBER({u_b}), ISNUMBER({u_a}), {u_b}={u_a}), 1, 0))")
                    row_m += 4
                
                ws_g[f'{get_column_letter(u_calc_start_col+1)}{r}'] = f"={'+'.join(pts_f) if pts_f else '0'}"
            
            calc_range = f"{get_column_letter(u_calc_start_col)}{calc_start_row+1}:{get_column_letter(u_calc_start_col+1)}{calc_start_row+4}"
            ws_g[f'{get_column_letter(c_u+1)}{calc_start_row+2}'] = f"=SORT({calc_range}, 2, FALSE)"
            
            for idx in range(4):
                ws_g[f'{start_col_letter}{calc_start_row+2+idx}'] = idx + 1
            
            c_u += 5

    # --- WYNIKI: formuly sumujace i SORT ---
    ws_main.column_dimensions['Z'].hidden = True
    ws_main.column_dimensions['AA'].hidden = True
    for u in range(8):
        row = u + 7
        ws_main[f'Z{row}'] = f"=A{10+u}"
        
        sum_parts = []
        pts_col = get_column_letter(14 + u*5 + 3)
        for g_name in group_names:
            sum_parts.append(f"SUM('{g_name.replace('Group', 'Grupa')}'!{pts_col}:{pts_col})")
            
        ws_main[f'AA{row}'] = "=" + "+".join(sum_parts)
        
    # Auto sorting array
    ws_main['D7'] = "=SORT(Z7:AA14, 2, FALSE)"
    
    # Static formatting for leaderboard positions
    for row in range(7, 15):
        ws_main[f'D{row}'].alignment = Alignment(horizontal="center")
        ws_main[f'E{row}'].alignment = Alignment(horizontal="center")
        
        if row == 7: # 1st
            fill = PatternFill("solid", fgColor="FFD700")
            font = Font(bold=True, size=16)
        elif row == 8: # 2nd
            fill = PatternFill("solid", fgColor="C0C0C0")
            font = Font(bold=True, size=14)
        elif row == 9: # 3rd
            fill = PatternFill("solid", fgColor="CD7F32")
            font = Font(bold=True, size=12)
        else: # 4-8
            fill = PatternFill("solid", fgColor="E0E0E0")
            font = Font(bold=False, size=11)
            
        ws_main[f'D{row}'].fill = fill
        ws_main[f'D{row}'].font = font
        ws_main[f'E{row}'].fill = fill
        ws_main[f'E{row}'].font = font


    wb.save("WorldCup2026_Typer.xlsx")
    print("Wygenerowano nową wersję pliku: WorldCup2026_Typer.xlsx")

if __name__ == "__main__":
    groups, matches = parse_schedule()
    create_excel_v2(groups, matches)
